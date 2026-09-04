from __future__ import annotations

import csv
import io
import sqlite3
import tempfile
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .models import Lead
from .net import normalize_url


FIELD_ALIASES = {
    "name": {"company", "company name", "company_name", "name", "企业名称", "公司", "公司名称"},
    "market": {"market", "location", "市场", "地区"},
    "company_type": {"type", "company type", "company_type", "category", "类型", "企业类型"},
    "contact_first_name": {"contact first name", "first name", "contact_first_name", "联系人名", "名字"},
    "contact_last_name": {"contact last name", "last name", "contact_last_name", "联系人姓", "姓氏"},
    "email": {"contact info", "contact email", "email", "e-mail", "邮箱", "联系邮箱"},
    "phone": {"phone number (if available)", "phone", "telephone", "tel", "电话", "电话号码"},
    "contact_phone": {"contact_phone", "contact phone", "contact phone number", "联系人电话", "联系人手机号"},
    "job_title": {"job_title", "job title", "title", "职位", "职务"},
    "website": {"website", "web site", "url", "domain", "company domain name", "官网", "网站", "域名"},
    "address": {"address", "street address", "地址"},
    "city": {"city", "城市"},
    "state": {"state", "state/region", "region", "州", "州/地区"},
    "country": {"country", "country/region", "国家", "国家/地区"},
    "employee_count": {"employee_count", "employee count", "number of employees", "employees", "员工数", "员工人数"},
    "signal": {"signal", "business signal", "业务信号"},
    "scale": {"scale", "company scale", "规模", "企业规模"},
    "score": {"score", "rating", "评分", "分数"},
}

MISSING_FIELDS = (
    "market", "company_type", "contact_first_name", "contact_last_name",
    "email", "phone", "contact_phone", "job_title", "website", "address",
    "city", "state", "country", "employee_count", "signal", "scale",
)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalized_row(headers: list[str], values: Iterable[Any]) -> dict[str, str]:
    raw = {header.casefold().strip(): _text(value) for header, value in zip(headers, values)}
    result: dict[str, str] = {}
    for field, aliases in FIELD_ALIASES.items():
        result[field] = next((raw[alias] for alias in aliases if raw.get(alias)), "")
    return result


def _lead_from_row(row: dict[str, str]) -> Lead | None:
    name = row.get("name", "").strip()
    if not name:
        return None
    try:
        score = max(0, min(120, int(float(row.get("score", "0") or 0))))
    except ValueError:
        score = 0
    return Lead(
        name=name,
        market=row.get("market", ""),
        company_type=row.get("company_type", ""),
        contact_first_name=row.get("contact_first_name", ""),
        contact_last_name=row.get("contact_last_name", ""),
        email=row.get("email", ""),
        phone=row.get("phone", ""),
        contact_phone=row.get("contact_phone", ""),
        job_title=row.get("job_title", ""),
        website=normalize_url(row.get("website", "")),
        address=row.get("address", ""),
        city=row.get("city", ""),
        state=row.get("state", "") or row.get("market", ""),
        country=row.get("country", ""),
        employee_count=row.get("employee_count", ""),
        signal=row.get("signal", ""),
        scale=row.get("scale", ""),
        score=score,
        source="Original database import",
    )


def _parse_csv(payload: bytes) -> list[Lead]:
    text = payload.decode("utf-8-sig")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = iter(reader)
    headers = [_text(value) for value in next(rows, [])]
    if not headers:
        return []
    leads: list[Lead] = []
    for values in rows:
        lead = _lead_from_row(_normalized_row(headers, values))
        if lead:
            leads.append(lead)
    return leads


def _parse_xlsx(payload: bytes) -> list[Lead]:
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [_text(value) for value in next(rows, [])]
        if not headers:
            return []
        leads: list[Lead] = []
        for values in rows:
            lead = _lead_from_row(_normalized_row(headers, values))
            if lead:
                leads.append(lead)
        return leads
    finally:
        workbook.close()


def _quoted_identifier(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def _parse_sqlite(payload: bytes, limit: int) -> list[Lead]:
    if not payload.startswith(b"SQLite format 3\x00"):
        raise ValueError("invalid SQLite database")
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as temporary:
            temporary.write(payload)
            temporary_path = Path(temporary.name)
        connection = sqlite3.connect(f"file:{temporary_path.as_posix()}?mode=ro", uri=True)
        try:
            connection.execute("PRAGMA query_only = ON")
            connection.execute("PRAGMA trusted_schema = OFF")
            tables = connection.execute("""
                SELECT name FROM sqlite_master
                WHERE type = 'table' AND name NOT LIKE 'sqlite_%'
                ORDER BY name
            """).fetchall()
            leads: list[Lead] = []
            for (table_name,) in tables:
                quoted_table = _quoted_identifier(table_name)
                columns = [
                    row[1] for row in connection.execute(
                        f"PRAGMA table_info({quoted_table})"
                    ).fetchall()
                ]
                normalized_columns = {column.casefold().strip() for column in columns}
                if not normalized_columns.intersection(FIELD_ALIASES["name"]):
                    continue
                remaining = limit - len(leads)
                if remaining <= 0:
                    break
                for values in connection.execute(
                    f"SELECT * FROM {quoted_table} LIMIT ?", (remaining,)
                ).fetchall():
                    lead = _lead_from_row(_normalized_row(columns, values))
                    if lead:
                        leads.append(lead)
            return leads
        finally:
            connection.close()
    finally:
        if temporary_path:
            temporary_path.unlink(missing_ok=True)


def parse_database_file(filename: str, payload: bytes, limit: int = 10_000) -> list[Lead]:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".csv":
        leads = _parse_csv(payload)
    elif suffix == ".xlsx":
        leads = _parse_xlsx(payload)
    elif suffix in {".db", ".sqlite", ".sqlite3"}:
        leads = _parse_sqlite(payload, limit)
    else:
        raise ValueError("unsupported database file type")
    return leads[:limit]


def missing_fields(lead: Lead) -> list[str]:
    return [field for field in MISSING_FIELDS if not getattr(lead, field).strip()]
